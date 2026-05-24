import json
from typing import TypedDict
from langchain_core.messages import SystemMessage, HumanMessage, ToolMessage
from langgraph.graph import StateGraph, END
from langchain_google_genai import ChatGoogleGenerativeAI
from src.agent.state import AgentState
from src.agent.router import route_file
from src.agent.prompts import SYSTEM_PROMPT

from langgraph.checkpoint.memory import MemorySaver

# Local extraction imports (zero API calls)
import fitz
import pdfplumber
import openpyxl


def _extract_pdf(file_path: str) -> str:
    """Extract text + metadata + tables from PDF. Zero API calls."""
    doc = fitz.open(file_path)
    
    # Metadata block
    meta = doc.metadata or {}
    meta_block = (
        f"File Metadata:\n"
        f"  Title: {meta.get('title','').strip() or 'N/A'}\n"
        f"  Author: {meta.get('author','').strip() or 'N/A'}\n"
        f"  Creator: {meta.get('creator','').strip() or 'N/A'}\n"
        f"  Pages: {len(doc)}\n"
    )
    
    parts = [meta_block]
    for i, page in enumerate(doc):
        # sort=True fixes multi-column PDFs by reading left→right, top→bottom
        text = page.get_text(sort=True).strip()
        if text:
            parts.append(f"--- Page {i+1} ---\n{text}")
    
    # Extract tables with pdfplumber
    try:
        with pdfplumber.open(file_path) as pdf:
            for i, page in enumerate(pdf.pages):
                for t_idx, table in enumerate(page.extract_tables()):
                    formatted = "\n".join(" | ".join(str(c) if c else "" for c in row) for row in table)
                    parts.append(f"--- Table {t_idx+1} on Page {i+1} ---\n{formatted}")
    except Exception:
        pass
    
    return "\n\n".join(parts)


def _extract_excel(file_path: str) -> str:
    """Extract content + column schema + data quality hints from Excel. Zero API calls."""
    import datetime
    wb = openpyxl.load_workbook(file_path, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue

        headers = [str(h) if h is not None else f"Col{i+1}" for i, h in enumerate(all_rows[0])]
        data_rows = all_rows[1:]

        # Column type inference
        col_types = []
        for col_idx, header in enumerate(headers):
            vals = [r[col_idx] if col_idx < len(r) else None for r in data_rows]
            non_null = [v for v in vals if v is not None and str(v).strip()]
            if not non_null:
                col_types.append(f"{header}: empty")
                continue
            type_counts = {"int": 0, "float": 0, "date": 0, "text": 0}
            for v in non_null:
                if isinstance(v, (datetime.datetime, datetime.date)):
                    type_counts["date"] += 1
                elif isinstance(v, int):
                    type_counts["int"] += 1
                elif isinstance(v, float):
                    type_counts["float"] += 1
                else:
                    try:
                        float(str(v).replace(",", ""))
                        type_counts["float"] += 1
                    except ValueError:
                        type_counts["text"] += 1
            dominant = max(type_counts, key=type_counts.get)
            ratio = type_counts[dominant] / len(non_null)
            typ = dominant if ratio >= 0.6 else "mixed"
            null_pct = round((len(vals) - len(non_null)) / len(vals) * 100) if vals else 0
            col_types.append(f"{header}: {typ}{f' ({null_pct}% empty)' if null_pct else ''}")

        # Duplicate detection
        seen = set()
        dupes = sum(1 for r in data_rows if (k := str(r)) and k in seen and any(v is not None for v in r) or not seen.add(k))
        # Correct duplicate count
        seen2 = set()
        dupes = 0
        for r in data_rows:
            k = str(r)
            if k in seen2 and any(v is not None for v in r):
                dupes += 1
            seen2.add(k)

        parts.append(f"--- Sheet: {sheet_name} ({ws.max_row} rows × {ws.max_column} cols) ---")
        parts.append(f"Column Schema: {', '.join(col_types)}")
        if dupes:
            parts.append(f"⚠️ {dupes} duplicate row(s) detected")
        parts.append("Headers: " + " | ".join(headers))
        for row in data_rows[:30]:  # Cap at 30 rows to save tokens
            parts.append(" | ".join(str(c) if c is not None else "" for c in row))
        if len(data_rows) > 30:
            parts.append(f"... ({len(data_rows) - 30} more rows truncated)")
    
    return "\n".join(parts)


def build_graph(skills_map=None, llm=None):
    if not llm:
        llm = ChatGoogleGenerativeAI(model="gemini-3-flash-preview")
    
    # Create the checkpointer
    memory = MemorySaver()

    def router_node(state: AgentState):
        """
        Node 1: Deterministic file routing + content extraction.
        Extracts content locally (zero API calls) and injects it
        directly into the message history for the LLM.
        
        On follow-up questions (document already in history), skips
        re-extraction entirely to avoid duplicate context and token waste.
        """
        file_path = state.get("file_path")
        if not file_path:
            return state  # Pure chat message, no file involved

        # Guard: if document is already extracted in this thread,
        # skip re-extraction. This makes multi-turn conversations efficient.
        existing_messages = list(state.get("messages", []))
        already_extracted = any(
            isinstance(m, HumanMessage)
            and isinstance(m.content, str)
            and m.content.startswith("[DOCUMENT CONTEXT]")
            for m in existing_messages
        )
        if already_extracted:
            return state  # Document already in history — go straight to agent

        routing_info = route_file(file_path)
        file_type = routing_info["file_type"]

        # --- LOCAL EXTRACTION (zero API calls) ---
        if "pdf" in file_type:
            content = _extract_pdf(file_path)
        elif "spreadsheet" in file_type or "excel" in file_type or "csv" in file_type:
            content = _extract_excel(file_path)
        else:
            content = "(Unsupported file type for extraction)"

        # Inject extracted content as a hidden context message
        context_msg = HumanMessage(
            content=(
                f"[DOCUMENT CONTEXT]\n"
                f"File: {file_path}\n"
                f"Type: {file_type}\n\n"
                f"{content}"
            )
        )
        return {
            "file_type": file_type,
            "messages": [context_msg]
        }

    def agent_node(state: AgentState):
        """
        Node 2: Single LLM call. 
        The document content is already extracted and sitting 
        in the message history. The LLM just needs to analyze 
        and produce the final formatted response.
        No tool calls needed = no looping.
        """
        messages = [SystemMessage(content=SYSTEM_PROMPT)] + list(state["messages"])
        response = llm.invoke(messages)
        return {"messages": [response]}

    # --- SIMPLIFIED GRAPH: router → agent → END ---
    # No tool loop needed = 1 API call total
    graph = StateGraph(AgentState)
    graph.add_node("router", router_node)
    graph.add_node("agent", agent_node)

    graph.set_entry_point("router")
    graph.add_edge("router", "agent")
    graph.add_edge("agent", END)

    return graph.compile(checkpointer=memory)
