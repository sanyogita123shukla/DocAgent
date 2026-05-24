import openpyxl
from typing import Any, Type
from pydantic import BaseModel
from src.skills.base import AbstractSkill, SkillInput

class ExcelReaderSkill(AbstractSkill):
    name: str = "excel_reader"
    description: str = "Reads an Excel file, handles merged cells and multiple sheets."
    args_schema: Type[BaseModel] = SkillInput

    def _resolve_merged(self, cell, sheet) -> Any:
        """
        Returns the value of a cell, resolving to the top-left value 
        if the cell is part of a merged range.
        """
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # the top-left cell has the value
                return sheet[merged_range.start_cell.coordinate].value
        return cell.value

    def _execute(self, file_path: str, **kwargs: Any) -> dict[str, Any]:
        """
        Extracts data from Excel, handling multiple sheets and merged cells.
        """
        # load_workbook(data_only=True) reads formula values instead of the formula string
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheets_data = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows():
                # Extract values, managing merged cells
                rows.append([self._resolve_merged(cell, ws) for cell in row])
            
            headers = rows[0] if rows else []
            data = rows[1:]
            
            sheets_data[sheet_name] = {
                "headers": headers,
                "data": data,
                "shape": (ws.max_row, ws.max_column),
            }
        return {"sheets": sheets_data, "sheet_names": wb.sheetnames}
