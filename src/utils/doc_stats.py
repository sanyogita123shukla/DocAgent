"""
Local document statistics — computed with zero API calls.

Priority 1 features added:
- PDF metadata (author, title, created date)
- PDF table count
- PDF multi-column reading order fix
- PDF hyperlink count
- Excel column type inference (int / float / date / text / mixed)
- Excel duplicate row detection
- Excel empty columns detection
- Local document intent heuristic (form / report / dataset / contract)
"""
import re
import datetime
import fitz
import openpyxl
import pdfplumber

# ─────────────────────────────────────────────────────────────────────
# Regex patterns for form/question detection
# ─────────────────────────────────────────────────────────────────────
QUESTION_PATTERNS = [
    r'\d+[\.\)]\s+.+\?',                    # "1. What is...?"
    r'(?:Q|Question)\s*\d*[:\.\)]\s*.+',    # "Q1: ..."
    r'^[A-Z][\.\)]\s+.+\?',                 # "A. What...?"
    r'_{3,}',                               # "___________ " fill-in blanks
    r'\[\s*\]',                             # "[ ]" checkboxes
    r'\[X\]',                               # "[X]" checked boxes
    r'(?:Yes|No)\s*/\s*(?:Yes|No)',         # "Yes / No"
    r'(?:Please|Kindly)\s+(?:describe|explain|list|provide)',
]

# Keywords for local intent heuristics
FORM_KEYWORDS = {"questionnaire", "form", "registration", "intake", "application", "consent", "survey"}
REPORT_KEYWORDS = {"summary", "report", "briefing", "analysis", "overview", "findings", "conclusion", "abstract"}
CONTRACT_KEYWORDS = {"agreement", "contract", "terms", "obligations", "clause", "party", "herein", "whereas"}
DATASET_KEYWORDS = {"dataset", "data", "records", "entries", "rows", "columns", "database", "spreadsheet"}


# ─────────────────────────────────────────────────────────────────────
# LOCAL INTENT HEURISTIC (zero API calls)
# ─────────────────────────────────────────────────────────────────────
def _classify_intent(text: str, question_count: int, file_type: str) -> str:
    """Classify document intent using keyword heuristics. Zero API calls."""
    text_lower = text.lower()
    
    if file_type == "Excel":
        return "Dataset"
    
    # Score each category
    scores = {
        "Form / Questionnaire": sum(1 for kw in FORM_KEYWORDS if kw in text_lower) + (2 if question_count >= 3 else 0),
        "Report / Briefing":    sum(1 for kw in REPORT_KEYWORDS if kw in text_lower),
        "Contract / Agreement": sum(1 for kw in CONTRACT_KEYWORDS if kw in text_lower),
        "Dataset / Spreadsheet": sum(1 for kw in DATASET_KEYWORDS if kw in text_lower),
    }
    
    best = max(scores, key=scores.get)
    return best if scores[best] > 0 else "General Document"


# ─────────────────────────────────────────────────────────────────────
# PDF STATS
# ─────────────────────────────────────────────────────────────────────
def compute_pdf_stats(file_path: str) -> dict:
    """Extract rich stats from a PDF file. Zero API calls."""
    doc = fitz.open(file_path)
    total_pages = len(doc)

    # Metadata
    meta = doc.metadata or {}
    author   = meta.get("author", "").strip() or "Unknown"
    title    = meta.get("title", "").strip() or "Untitled"
    creator  = meta.get("creator", "").strip() or ""
    raw_date = meta.get("creationDate", "")
    # Parse PDF date format: D:YYYYMMDDHHmmSS
    created_date = "Unknown"
    if raw_date and raw_date.startswith("D:"):
        try:
            created_date = datetime.datetime.strptime(raw_date[2:10], "%Y%m%d").strftime("%d %b %Y")
        except Exception:
            pass

    # Text extraction — use sort=True to fix multi-column reading order
    all_text = ""
    hyperlink_count = 0
    for page in doc:
        all_text += page.get_text(sort=True) + "\n"   # sort=True = proper reading order
        hyperlink_count += len(page.get_links())

    lines = [l for l in all_text.split("\n") if l.strip()]
    words = all_text.split()

    # Detect form questions/fields via regex
    detected_questions = []
    for line in lines:
        for pattern in QUESTION_PATTERNS:
            if re.search(pattern, line.strip(), flags=re.IGNORECASE):
                detected_questions.append(line.strip())
                break

    # Embedded image count
    image_count = sum(len(page.get_images(full=False)) for page in doc)

    # Table count via pdfplumber
    table_count = 0
    try:
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                table_count += len(page.extract_tables())
    except Exception:
        pass

    intent = _classify_intent(all_text, len(detected_questions), "PDF")

    return {
        "file_type": "PDF",
        "intent": intent,
        # Volume
        "pages": total_pages,
        "words": len(words),
        "lines": len(lines),
        "characters": len(all_text),
        # Structure
        "tables": table_count,
        "images": image_count,
        "hyperlinks": hyperlink_count,
        "questions_detected": len(detected_questions),
        # Metadata
        "meta_title": title,
        "meta_author": author,
        "meta_creator": creator,
        "meta_created": created_date,
    }


# ─────────────────────────────────────────────────────────────────────
# EXCEL STATS
# ─────────────────────────────────────────────────────────────────────
def _infer_col_type(values: list) -> str:
    """Infer the dominant data type of a column. Zero API calls."""
    non_null = [v for v in values if v is not None and str(v).strip() != ""]
    if not non_null:
        return "empty"
    
    type_counts = {"integer": 0, "float": 0, "date": 0, "text": 0}
    for v in non_null:
        if isinstance(v, datetime.datetime) or isinstance(v, datetime.date):
            type_counts["date"] += 1
        elif isinstance(v, int):
            type_counts["integer"] += 1
        elif isinstance(v, float):
            type_counts["float"] += 1
        else:
            # Try numeric conversion
            try:
                float(str(v).replace(",", ""))
                type_counts["float"] += 1
            except ValueError:
                type_counts["text"] += 1

    dominant = max(type_counts, key=type_counts.get)
    total = len(non_null)
    dominance = type_counts[dominant] / total

    if dominance < 0.6:
        return "mixed"
    return dominant


def compute_excel_stats(file_path: str) -> dict:
    """Extract rich stats from an Excel file. Zero API calls."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    total_sheets = len(wb.sheetnames)
    total_rows = 0
    total_cells = 0
    total_filled_cells = 0
    column_types_summary = {}
    duplicate_rows_total = 0
    empty_columns_total = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue

        total_rows += ws.max_row
        headers = [str(h) if h is not None else f"Col{i+1}" for i, h in enumerate(all_rows[0])]
        data_rows = all_rows[1:]

        # Cell counts
        for row in all_rows:
            for cell in row:
                total_cells += 1
                if cell is not None and str(cell).strip():
                    total_filled_cells += 1

        # Column type inference
        col_types = {}
        for col_idx, header in enumerate(headers):
            col_values = [row[col_idx] if col_idx < len(row) else None for row in data_rows]
            inferred = _infer_col_type(col_values)
            col_types[header] = inferred
            if inferred == "empty":
                empty_columns_total += 1
        column_types_summary[sheet_name] = col_types

        # Duplicate row detection (compare stringified rows)
        if data_rows:
            seen = set()
            dupes = 0
            for row in data_rows:
                key = str(row)
                if key in seen and any(v is not None for v in row):
                    dupes += 1
                seen.add(key)
            duplicate_rows_total += dupes

    intent = _classify_intent("", 0, "Excel")

    return {
        "file_type": "Excel",
        "intent": intent,
        # Volume
        "sheets": total_sheets,
        "sheet_names": wb.sheetnames,
        "rows": total_rows,
        "cells_total": total_cells,
        "cells_filled": total_filled_cells,
        # Quality
        "duplicate_rows": duplicate_rows_total,
        "empty_columns": empty_columns_total,
        "fill_rate": round(total_filled_cells / total_cells * 100, 1) if total_cells > 0 else 0,
        # Column schema
        "column_types": column_types_summary,
    }


# ─────────────────────────────────────────────────────────────────────
# DISPATCHER
# ─────────────────────────────────────────────────────────────────────
def compute_stats(file_path: str) -> dict | None:
    """Auto-detect file type and compute stats. Returns None if unsupported."""
    lower = file_path.lower()
    if lower.endswith(".pdf"):
        return compute_pdf_stats(file_path)
    elif lower.endswith((".xlsx", ".xls", ".csv")):
        return compute_excel_stats(file_path)
    return None
